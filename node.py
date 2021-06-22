from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.cell import WriteOnlyCell
from openpyxl import load_workbook
import pandas as pd

#tree structure
class Node:
    def __init__(self, height, name, number, info):
        self.height = int(height)
        self.name = name
        self.number = int(number)
        self.info = info
        self.children = []
        self.find = 0
        self.iid = -1
    def addchildren(self, child) :
        self.children.append(child)
    def equal(self, other) :
        if self.name == other.name :
            return True
        return False
    def numberEq (self, other):
        if self.number == other.number :
            return True
        return False
red = Font(color="00FF0000")
#write treenode 1, actually used only for unmatch
def writenode1(node, isred, ws):
    if isred:
        h = WriteOnlyCell(ws, value=node.height)
        h.font = red
        n = WriteOnlyCell(ws, value=node.name)
        n.font = red
        num = WriteOnlyCell(ws, value=node.number)
        num.font = red
        i = WriteOnlyCell(ws, value=node.info)
        i.font = red
        ws.append([h, n, num, i])
    else:
        ws.append([node.height, node.name, node.number, node.info])
#write treenode 2, actually used only for unmatch
def writenode2(node, isred, ws):
    if isred:
        h = WriteOnlyCell(ws, value=node.height)
        h.font = red
        n = WriteOnlyCell(ws, value=node.name)
        n.font = red
        num = WriteOnlyCell(ws, value=node.number)
        num.font = red
        i = WriteOnlyCell(ws, value=node.info)
        i.font = red
        ws.append(["","","","","",h, n, num, i])
    else:
        ws.append(["","","","","",node.height, node.name, node.number, node.info])
def dfs(node) :
    print(node.name)
    if node.children == []:
        return
    for i in node.children:
        dfs(i)
#for form 2 test, dictionary tree
def dfs2(dic, parent):
    if dic.get(parent.name) != None:
        print(parent.name)
        print("contains:")
        children = dic[parent.name]
        for child in children:
            print(child.name)
        for child in children:
            dfs2(dic, child)
def buildtree(dic, parent):
    if dic.get(parent.name) != None:
        children = dic[parent.name]
        for child in children:
            child.height = parent.height + 1
            parent.children.append(child)
        for child in children:
            buildtree(dic, child)
def buildform(node, ws):
    ws.append([node.height, node.name, node.number, node.info])
    for i in node.children:
        buildform(i, ws)
#declare an empty node, used for node match, dfs a single tree
empty = Node(-1, "", -1, "")
#compare two tree. unmatch node is marked red
def compare(node1, node2, ws, bound, titles):
    if node1.children == [] and node2.children == []:
        return
    elif node1.children == []:
        if node2.height == bound:
            return
        for i in node2.children:
            if i.name[-1] == "*":
                i.name = i.name[:-1]
            writenode2(i, True, ws)
            compare(empty, i, ws, bound)
        return
    elif node2.children == []:
        if node1.height == bound:
            return
        for i in node1.children:
            writenode1(i, True, ws)
            compare(i, empty, ws, bound)  
        return
    if node1.height == bound:
        return
    dic = {}
    for n1 in node1.children:
        dic[n1.name] = n1
    for n2 in node2.children:
        if n2.name[-1] == "*":
            n2.name = n2.name[:-1]
        if dic.get(n2.name) != None:
            n1 = dic.get(n2.name)
            ws.append([n1.height, n1.name, n1.number, n1.info, "", n2.height, n2.name, n2.number, n2.info])
            n1.find = 1
            n2.find = 1
            compare(n1, n2, ws, bound)
    for n1 in node1.children:
        if n1.find == 0:
            writenode1(n1, True, ws)
            compare(n1, empty, ws, bound)
    for n2 in node2.children:
        if n2.find == 0:
            writenode2(n2, True, ws)
            compare(empty, n2, ws, bound)

def form1(filename):
    hierachy1 = {}
    count = 0
    root =  Node(0,"default",0,"default")
    #convert form 1 into tree
    if filename[-4:] == "xlsx":
        wb = load_workbook(filename, read_only=True)
        wsname = wb.sheetnames
        ws = wb[wsname[0]]
        for line in ws.rows:
            if count == 0:
                count += 1
                continue;
            height = line[0].value
            if height == None:
                break
            name = line[1].value.strip()
            number = line[2].value
            project = line[3].value
            sign = line[4].value
            condition = line[5].value
            info = project + "  " + str(sign) + "  " + condition
            node = Node(height, name, number, info)
            height = int(height)
            if height == 0:
                root = node
            hierachy1[height] = node
            parent = hierachy1.get(height - 1)
            if parent != None:
                parent.children.append(node)
    else :
        f = open(filename, 'r')
        for line in f:
            if count == 0:
                count += 1
                continue;
            firstindex = line.find("\t")
            height = line[:firstindex];
            if height == "":
                break
            line = line[firstindex:];
            line = line.lstrip();
            secondindex = line.find("\t")
            name = line[0 : secondindex]

            line = line[secondindex:];
            line = line.lstrip();
            thirdindex = line.find("\t")
            number = line[0 : thirdindex]

            line = line[thirdindex:];
            line = line.lstrip();
            index = line.find("\t")
            project = line[0 : index]

            line = line[index:];
            line = line.lstrip();
            index = line.find("\t")
            sign = line[0 : index]

            line = line[index:];
            line = line.lstrip();
            condition = line.find("\t")
            condition = line[0 : index]

            info = project + "\t" + sign + "\t" + condition
            node = Node(height, name, number, info)
            height = int(height)
            if height == 0:
                root = node
            hierachy1[height] = node
            parent = hierachy1.get(height - 1)
            if parent != None:
                parent.children.append(node)
        f.close()
    # wb1 = Workbook()
    # ws1 = wb1.active
    # buildform(root, ws1)
    # wb1.save("output/form1.xlsx")
    return root
    
    # page.f1 = 1

    # dfs(root1)
def form2(filename):
    data2 = {}
    rootname = ""
    parentname = ""
    if filename[-4:] == "xlsx":
        wb = load_workbook(filename, read_only=True)
        wsname = wb.sheetnames
        ws = wb[wsname[0]]
        for line in ws.rows:
            if line[0].value != None:
                title = line[0].value
                if (title == "装配零件的总数"):
                    break
                parentname = line[1].value.strip()
                data2[parentname] = []
                if rootname == "":
                    rootname = parentname
            else:
                number = line[1].value
                if number == None:
                    continue
                info = line[2].value
                name = line[3].value.strip()
                if (info == "零件"):
                    name = name + "*"
                node = Node(-1, name, number, info)
                data2[parentname].append(node)
        wb.close()
    else:
        f = open(filename, 'r')
        for line in f:
            firstindex = line.find("\t")

            if firstindex != 0:
                title = line[:firstindex]
                if (title == "装配零件的总数"):
                    break
                line = line[firstindex:]
                line = line.lstrip();
                secondindex = line.find("\t")
                parentname = line[0 : secondindex]
                data2[parentname] = []
                if rootname == "":
                    rootname = parentname
            else:
                line = line.lstrip();
                secondindex = line.find("\t")
                number = line[0 : secondindex]
                if number == "":
                    continue
                line = line[secondindex:]
                line = line.lstrip();
                secondindex = line.find("\t")
                info = line[0 : secondindex]

                line = line[secondindex:]
                name = line.strip();
                if (info == "零件"):
                    name = name + "*"
                node = Node(-1, name, number, info)
                data2[parentname].append(node)
        f.close()
    root = Node(0, rootname, 1, "")
    #tree two into tree
    buildtree(data2, root)
    # wb2 = Workbook()
    # ws2 = wb2.active
    # buildform(root, ws2)
    # wb2.save("output/form2.xlsx")
    return root
    

